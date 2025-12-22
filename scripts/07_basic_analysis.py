#!/usr/bin/env python3
"""
Run a full metabarcoding analysis across **all** ASVs in a filtered count table.

This script is a corrected version of the original `run_full_metabarcoding_analysis.py`.  The
major changes address the issue where metadata columns (e.g., latitude/longitude,
country, BIN codes, etc.) were being interpreted as sample read counts.  To
prevent this, the script now uses both an explicit list of known metadata
columns and a regular expression to identify legitimate sample columns.  Only
columns matching the sample naming pattern (e.g. `suh_a/1`, `kan_b/3/20`) and
not in the metadata list are treated as samples.

Features:
  * Computes per‑sample alpha diversity metrics (library size, observed ASVs,
    Shannon index, Pielou evenness) on all ASV rows.
  * Counts unique species, genera and families across all samples.
  * Aggregates read counts by order and genus, then computes relative
    composition per sample and per site.
  * Performs non‑metric multidimensional scaling (NMDS) on genus‑level
    Bray–Curtis distances when scipy and sklearn are available; otherwise
    gracefully skips.
  * Writes results to an Excel workbook with multiple sheets, along with
    several PNG figures and a summary JSON file.  The Excel writer uses
    openpyxl by default, removing any dependency on XlsxWriter.

Usage:
    python run_full_metabarcoding_analysis_corrected.py \
        --input filtered_ASV_table.xlsx \
        --outdir results_full

Dependencies:
    - openpyxl, pandas, numpy, matplotlib
    - Optionally scipy and scikit‑learn for NMDS

Author: OpenAI's ChatGPT (corrected version)
"""
import argparse
import sys  # Needed for printing progress to stderr
import json
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
from matplotlib import cm
from openpyxl import load_workbook

# Default locations so the script can be run without flags
DEFAULT_INPUT = "/mnt/c/Users/vidna/Documents/mtb/data/mtb_forest_PHK/bioinfo/processed_data/filtered_output/filtered_ASV_table.xlsx"
DEFAULT_OUTDIR = "/mnt/c/Users/vidna/Documents/mtb/data/mtb_forest_PHK/bioinfo/processed_data/basic_analysis"


def pick_tax_col(col_idx, base):
    """Return the first matching taxonomy column name for a given base rank.

    Some tables may use variants like `order_x`, `order_y`.  This helper
    checks for the presence of such columns in priority order.
    """
    for cand in (f"{base}_x", base, f"{base}_y"):
        if cand in col_idx:
            return cand
    return None


def parse_sample_name(name: str) -> dict:
    """Parse a sample identifier into site, plot, replicate and depth.

    A sample name is expected to look like `site_plot/rep` (e.g., `kan_a/3`) or
    possibly `site_plot/rep/20` when indicating a deeper (10–20 cm) soil layer.
    Depth is flagged as `10-20` if the name contains `10-20`, `20cm`,
    `bottom`, `deep` or `bott` (case insensitive), otherwise it defaults to
    `0-10`.  Missing fields are set to empty strings.
    """
    s = str(name)
    site = plot = rep = ""
    depth = "0-10"
    lower = s.lower()
    if any(k in lower for k in ("10-20", "20cm", "bottom", "deep", "bott")):
        depth = "10-20"
    # Split on slash first
    if "/" in s:
        left, rep = s.split("/", 1)
        rep = rep.strip()
    else:
        left = s
    # Now split left part on underscore to get site/plot
    if "_" in left:
        site, plot = left.split("_", 1)
    else:
        site = left
        plot = ""
    return {"sample": s, "site": site, "plot": plot, "rep": rep, "depth": depth}

def get_treatment_metadata() -> dict:
    """Hard-coded treatment metadata (from provided table)."""
    # Normalize into treatment code + short and long descriptions
    rows = [
        ("B1_A", "A20", "20t_top", "20 tons ash/ha applied on top of the soil", 0.521, 14.8, 22.7),
        ("B1_C", "C", "control", "No addition or incorporation", 0.527, None, 29.6),
        ("B1_F", "A6", "6t_top", "6 tons ash/ha applied on top of the soil", 0.482, None, 41.2),
        ("B1_ART", "A20RT", "20t_rt", "20 tons ash/ha rotary tilled into the soil", 0.502, 5, 23.4),
        ("B1_RT", "RT", "rt_only", "No addition but soil is rotary tilled.", 0.51, None, 3.88),
        ("B2_A", "A20", "20t_top", "20 tons ash/ha applied on top of the soil", 0.518, 16.7, 56),
        ("B2_C", "C", "control", "No addition or incorporation", 0.521, None, 1.71),
        ("B2_F", "A6", "6t_top", "6 tons ash/ha applied on top of the soil", 0.522, None, 0.558),
        ("B2_ART", "A20RT", "20t_rt", "20 tons ash/ha rotary tilled into the soil", 0.507, None, 44.5),
        ("B2_RT", "RT", "rt_only", "No addition but soil is rotary tilled.", 0.514, None, 27.8),
        ("B3_A", "A20", "20t_top", "20 tons ash/ha applied on top of the soil", 0.52, 32, 40),
        ("B3_C", "C", "control", "No addition or incorporation", 0.535, None, 2.87),
        ("B3_F", "A6", "6t_top", "6 tons ash/ha applied on top of the soil", 0.523, None, 26.6),
        ("B3_ART", "A20RT", "20t_rt", "20 tons ash/ha rotary tilled into the soil", 0.511, None, 51),
        ("B3_RT", "RT", "rt_only", "No addition but soil is rotary tilled.", 0.53, None, 9.43),
        ("B4_A", "A20", "20t_top", "20 tons ash/ha applied on top of the soil", 0.482, 24.2, 41.2),
        ("B4_C", "C", "control", "No addition or incorporation", 0.529, 2.2, 50),
        ("B4_F", "A6", "6t_top", "6 tons ash/ha applied on top of the soil", 0.488, 6.86, 39),
        ("B4_ART", "A20RT", "20t_rt", "20 tons ash/ha rotary tilled into the soil", 0.488, 10, 46.3),
        ("B4_RT", "RT", "rt_only", "No addition but soil is rotary tilled.", 0.515, 5.43, 18.5),
        ("Skov_1A", "Skov", "ref", "Reference site", 0.26, 3.08, 37.2),
        ("Skov_1B", "Skov", "ref", "Reference site", 0.245, None, 1.36),
        ("Skov_2A", "Skov", "ref", "Reference site", 0.535, None, 23),
        ("Skov_3A", "Skov", "ref", "Reference site", 0.455, None, 0.906),
    ]
    meta = {}
    for sample, treat_code, treat_short, desc_long, dna, m1, m2 in rows:
        meta[sample] = {
            "treatment_code": treat_code,
            "treatment": treat_short,  # short label for grouping/plots
            "treatment_desc": desc_long,  # long description
            "dna_conc": dna,
            "metric1": m1,
            "metric2": m2,
        }
    return meta

def determine_sample_columns(header: list) -> list:
    """Return a list of column names that correspond to samples.

    Primary rule: everything before the first taxonomy column (e.g., phylum_x)
    is treated as sample data. This handles renamed samples like 'B1 A' that
    don't match the older regex.
    """
    tax_markers = {
        "kingdom", "kingdom_x", "kingdom_y",
        "phylum", "phylum_x", "phylum_y",
        "class", "class_x", "class_y",
        "order", "order_x", "order_y",
        "family", "family_x", "family_y",
        "genus", "genus_x", "genus_y",
        "species", "species_x", "species_y",
    }
    lower_header = [str(c).lower() for c in header]
    first_tax_idx = None
    for idx, name in enumerate(lower_header):
        if name in tax_markers:
            first_tax_idx = idx
            break
    if first_tax_idx is not None and first_tax_idx > 0:
        return header[:first_tax_idx]

    # Fallback: keep columns that look like sample names (legacy pattern)
    sample_regex = re.compile(r"^[A-Za-z]{2,3}_[A-Za-z]+/[0-9]+(?:[-/][0-9]+)?$", re.IGNORECASE)
    return [c for c in header if sample_regex.match(str(c))]


def main(inp: str, outdir: str) -> None:
    """Perform the metabarcoding analysis and write outputs to `outdir`."""
    if not os.path.isfile(inp):
        raise SystemExit(f"❌ Input file not found: {inp}\nProvide --input or update DEFAULT_INPUT in the script.")

    out_path = Path(outdir)
    out_path.mkdir(parents=True, exist_ok=True)
    # If the chosen outdir is not writable (e.g., locked by Windows), fall back to a local folder
    try:
        test_file = out_path / "_write_test.tmp"
        test_file.write_text("ok", encoding="utf-8")
        test_file.unlink(missing_ok=True)
    except PermissionError:
        fallback = Path.cwd() / "basic_analysis_output"
        fallback.mkdir(parents=True, exist_ok=True)
        print(f"⚠️ Output directory {out_path} not writable; switching to {fallback}")
        out_path = fallback

    # Load the workbook in read‑only streaming mode
    wb = load_workbook(inp, read_only=True, data_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(header)}

    # Determine taxonomy rank columns (some may be missing)
    tax_rank_cols = {rank: pick_tax_col(col_idx, rank) for rank in ["phylum", "class", "order", "family", "genus", "species"]}
    rank_indices = {rank: (col_idx[col] if col is not None else None) for rank, col in tax_rank_cols.items()}

    # Determine which columns are samples
    sample_cols = determine_sample_columns(header)
    sample_indices = [col_idx[c] for c in sample_cols]
    # Parse metadata from sample names
    meta_df = pd.DataFrame([parse_sample_name(c) for c in sample_cols])
    # Attach treatment metadata if available
    treat_meta = get_treatment_metadata()
    meta_df["sample_norm"] = meta_df["sample"].str.replace(" ", "_")
    meta_df["treatment"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("treatment"))
    meta_df["treatment_code"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("treatment_code"))
    meta_df["treatment_desc"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("treatment_desc"))
    meta_df["dna_conc"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("dna_conc"))
    meta_df["metric1"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("metric1"))
    meta_df["metric2"] = meta_df["sample_norm"].map(lambda x: treat_meta.get(x, {}).get("metric2"))

    # Accumulators for alpha and composition statistics
    libsize = Counter()
    richness = Counter()
    sum_c_logc = Counter()
    species_sets: dict[str, set] = defaultdict(set)
    genus_sets: dict[str, set] = defaultdict(set)
    family_sets: dict[str, set] = defaultdict(set)
    agg_by_order: dict[str, Counter] = defaultdict(Counter)
    agg_by_genus: dict[str, Counter] = defaultdict(Counter)
    agg_by_family: dict[str, Counter] = defaultdict(Counter)

    # Collect ASV-level counts for ASV-level beta/NMDS
    asv_counts = []  # list of lists in sample order

    # Iterate through ASV rows one by one to reduce memory usage
    total_rows = ws.max_row
    processed = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_val = row[rank_indices["order"]] if rank_indices["order"] is not None else None
        family_val = row[rank_indices["family"]] if rank_indices["family"] is not None else None
        genus_val = row[rank_indices["genus"]] if rank_indices["genus"] is not None else None
        species_val = row[rank_indices["species"]] if rank_indices["species"] is not None else None
        row_counts = []
        row_sum = 0.0
        for s_idx, sname in zip(sample_indices, sample_cols):
            val = row[s_idx]
            # Convert value to float if possible; missing or non‑numeric values count as zero
            if val is None or (isinstance(val, str) and not val.strip()):
                cnt = 0.0
            else:
                try:
                    cnt = float(val)
                except Exception:
                    cnt = 0.0
            row_counts.append(cnt)
            row_sum += cnt
            if cnt > 0.0:
                libsize[sname] += cnt
                richness[sname] += 1
                sum_c_logc[sname] += cnt * math.log(cnt)
                # Taxa sets
                if species_val not in (None, "", "NA"):
                    species_sets[sname].add(str(species_val))
                if genus_val not in (None, "", "NA"):
                    genus_sets[sname].add(str(genus_val))
                    agg_by_genus[sname][str(genus_val)] += cnt
                if family_val not in (None, "", "NA"):
                    family_sets[sname].add(str(family_val))
                    agg_by_family[sname][str(family_val)] += cnt
                if order_val not in (None, "", "NA"):
                    agg_by_order[sname][str(order_val)] += cnt
        if row_sum > 0:
            asv_counts.append(row_counts)
        processed += 1
        if processed % 10000 == 0:
            print(f"Processed {processed} rows / ~{total_rows - 1} (ASVs).", file=sys.stderr)

    # Construct alpha diversity DataFrame
    samples = sample_cols
    library_sizes = pd.Series([libsize[s] for s in samples], index=samples, name="library_size")
    observed_asvs = pd.Series([richness[s] for s in samples], index=samples, name="observed_asvs")
    shannon_values = []
    for s in samples:
        N = libsize[s]
        if N > 0 and observed_asvs[s] > 0:
            H_i = math.log(N) - (sum_c_logc[s] / N)
        else:
            H_i = 0.0
        shannon_values.append(H_i)
    shannon = pd.Series(shannon_values, index=samples, name="shannon")
    # Pielou evenness: H / log(observed ASVs)
    evenness = pd.Series([
        (shannon[s] / math.log(observed_asvs[s])) if observed_asvs[s] > 1 else np.nan
        for s in samples
    ], index=samples, name="pielou_evenness")
    species_rich = pd.Series([len(species_sets[s]) for s in samples], index=samples, name="species_richness")
    genus_rich = pd.Series([len(genus_sets[s]) for s in samples], index=samples, name="genus_richness")
    family_rich = pd.Series([len(family_sets[s]) for s in samples], index=samples, name="family_richness")
    alpha_df = pd.concat([
        library_sizes, observed_asvs, shannon, evenness, species_rich, genus_rich, family_rich
    ], axis=1).reset_index().rename(columns={"index": "sample"})
    alpha_annot = alpha_df.merge(meta_df, on="sample", how="left")

    # Overall counts of unique taxa across all samples
    overall_species = set().union(*species_sets.values()) if species_sets else set()
    overall_genera = set().union(*genus_sets.values()) if genus_sets else set()
    overall_families = set().union(*family_sets.values()) if family_sets else set()
    overall_counts = pd.DataFrame({
        "rank": ["species", "genus", "family"],
        "unique_taxa": [len(overall_species), len(overall_genera), len(overall_families)]
    })

    # Composition by order: relative abundances
    orders = sorted(set(k for s in agg_by_order.values() for k in s))
    comp_order = pd.DataFrame(0.0, index=samples, columns=orders)
    for s in samples:
        total = float(sum(agg_by_order[s].values()))
        if total > 0:
            for od, cnt in agg_by_order[s].items():
                comp_order.loc[s, od] = cnt / total
    comp_order = comp_order.reset_index().rename(columns={"index": "sample"})
    comp_order = comp_order.merge(meta_df, on="sample", how="left")

    # Composition by genus: relative abundances
    genera = sorted(set(k for s in agg_by_genus.values() for k in s))
    comp_genus = pd.DataFrame(0.0, index=samples, columns=genera)
    for s in samples:
        total = float(sum(agg_by_genus[s].values()))
        if total > 0:
            for g, cnt in agg_by_genus[s].items():
                comp_genus.loc[s, g] = cnt / total
    comp_genus = comp_genus.reset_index().rename(columns={"index": "sample"})
    comp_genus = comp_genus.merge(meta_df, on="sample", how="left")

    # Composition by family: relative abundances
    families = sorted(set(k for s in agg_by_family.values() for k in s))
    comp_family = pd.DataFrame(0.0, index=samples, columns=families)
    for s in samples:
        total = float(sum(agg_by_family[s].values()))
        if total > 0:
            for fam, cnt in agg_by_family[s].items():
                comp_family.loc[s, fam] = cnt / total
    comp_family = comp_family.reset_index().rename(columns={"index": "sample"})
    comp_family = comp_family.merge(meta_df, on="sample", how="left")
    # Aggregated compositions by treatment (exclude metadata columns)
    meta_cols = {"sample", "site", "plot", "rep", "depth", "sample_norm", "treatment", "treatment_code", "treatment_desc", "dna_conc", "metric1", "metric2"}
    if not comp_order.empty:
        order_tax_cols = [c for c in comp_order.columns if c not in meta_cols]
        comp_order_treat = comp_order[["treatment"] + order_tax_cols].groupby("treatment", as_index=False).mean(numeric_only=True)
    else:
        comp_order_treat = pd.DataFrame()
    if not comp_family.empty:
        family_tax_cols = [c for c in comp_family.columns if c not in meta_cols]
        comp_family_treat = comp_family[["treatment"] + family_tax_cols].groupby("treatment", as_index=False).mean(numeric_only=True)
    else:
        comp_family_treat = pd.DataFrame()

    # Beta diversity on genus relative abundances (Bray–Curtis and Jaccard presence/absence)
    genus_rel = comp_genus.set_index("sample")[genera]
    genus_presence = (genus_rel > 0).astype(float)

    def _pairwise(df_vals, idx, dist_func):
        n = len(idx)
        mat = np.zeros((n, n), dtype=float)
        for i in range(n):
            for j in range(i + 1, n):
                d = dist_func(df_vals[i], df_vals[j])
                mat[i, j] = mat[j, i] = d
        return pd.DataFrame(mat, index=idx, columns=idx)

    def _bray(u, v):
        denom = u.sum() + v.sum()
        if denom == 0:
            return 0.0
        return np.abs(u - v).sum() / denom

    def _jaccard(u, v):
        inter = np.logical_and(u > 0, v > 0).sum()
        union = np.logical_or(u > 0, v > 0).sum()
        return 1.0 - (inter / union) if union > 0 else 0.0

    bc_df = _pairwise(genus_rel.values, genus_rel.index.tolist(), _bray)
    jaccard_df = _pairwise(genus_presence.values, genus_presence.index.tolist(), _jaccard)

    def summarize_within_between(dist_df: pd.DataFrame, meta: pd.DataFrame, group_col: str = "site") -> pd.DataFrame:
        """Return long-form distances labeled within/between a grouping variable."""
        records = []
        group_map = meta.set_index("sample")[group_col].to_dict() if group_col in meta.columns else {}
        samples_idx = list(dist_df.index)
        for i, s1 in enumerate(samples_idx):
            for j in range(i + 1, len(samples_idx)):
                s2 = samples_idx[j]
                d = dist_df.iloc[i, j]
                if pd.isna(d):
                    continue
                g1 = group_map.get(s1)
                g2 = group_map.get(s2)
                comp = "within" if g1 is not None and g1 == g2 else "between"
                records.append({"sample1": s1, "sample2": s2, "comparison": comp, "distance": float(d)})
        return pd.DataFrame(records)

    dist_genus_long = summarize_within_between(bc_df, meta_df, group_col="treatment") if not bc_df.empty else pd.DataFrame()

    def compare_within_between(dist_long: pd.DataFrame, label: str) -> dict:
        """Return summary stats and Mann-Whitney p-value for within vs between."""
        if dist_long.empty:
            return {"level": label, "within_median": np.nan, "between_median": np.nan, "pvalue": np.nan}
        within = dist_long.loc[dist_long["comparison"] == "within", "distance"]
        between = dist_long.loc[dist_long["comparison"] == "between", "distance"]
        pval = np.nan
        try:
            from scipy.stats import mannwhitneyu
            if len(within) > 0 and len(between) > 0:
                pval = mannwhitneyu(within, between, alternative="two-sided").pvalue
        except Exception:
            pval = np.nan
        return {
            "level": label,
            "within_median": float(np.median(within)) if len(within) else np.nan,
            "between_median": float(np.median(between)) if len(between) else np.nan,
            "pvalue": pval,
        }

    def run_permanova(dist_df: pd.DataFrame, meta: pd.DataFrame, group_col: str = "site") -> dict:
        """Run PERMANOVA on a distance matrix by group (if skbio is available)."""
        if dist_df.empty or group_col not in meta.columns:
            return {"method": "permanova", "group": group_col, "pvalue": np.nan, "pseudoF": np.nan}
        try:
            from skbio.stats.distance import DistanceMatrix
            from skbio.stats.distance import permanova
            groups = meta.set_index("sample")[group_col].reindex(dist_df.index)
            dm = DistanceMatrix(dist_df.values, ids=dist_df.index)
            # skbio permanova uses positional grouping argument (named 'grouping' in >=0.5.8)
            res = permanova(dm, groups, permutations=999)
            return {"method": "permanova", "group": group_col, "pvalue": float(res["p-value"]), "pseudoF": float(res["test statistic"])}
        except Exception:
            # Fallback PERMANOVA implementation (Anderson 2001) using squared distances
            groups = meta.set_index("sample")[group_col].reindex(dist_df.index)
            valid = groups.notna()
            groups = groups[valid]
            if len(groups) < 3 or groups.nunique() < 2:
                return {"method": "permanova", "group": group_col, "pvalue": np.nan, "pseudoF": np.nan}
            dm = dist_df.loc[groups.index, groups.index].values
            n = dm.shape[0]
            # Precompute upper triangle squared distances
            triu_idx = np.triu_indices(n, k=1)
            d2 = dm[triu_idx] ** 2
            total_ss = d2.sum() / n
            unique_groups = groups.unique()
            within_ss = 0.0
            for g in unique_groups:
                idx = np.where(groups.values == g)[0]
                if len(idx) < 2:
                    continue
                tri = np.triu_indices(len(idx), k=1)
                g_d2 = dm[np.ix_(idx, idx)][tri] ** 2
                within_ss += g_d2.sum() / len(idx)
            between_ss = total_ss - within_ss
            df1 = len(unique_groups) - 1
            df2 = n - len(unique_groups)
            pseudoF = np.nan
            if df1 > 0 and df2 > 0 and within_ss > 0:
                pseudoF = (between_ss / df1) / (within_ss / df2)
            # Permutation test
            rng = np.random.default_rng(42)
            perms = 999
            perm_F = []
            for _ in range(perms):
                perm_labels = groups.values.copy()
                rng.shuffle(perm_labels)
                perm_within = 0.0
                for g in unique_groups:
                    idx = np.where(perm_labels == g)[0]
                    if len(idx) < 2:
                        continue
                    tri = np.triu_indices(len(idx), k=1)
                    g_d2 = dm[np.ix_(idx, idx)][tri] ** 2
                    perm_within += g_d2.sum() / len(idx)
                perm_between = total_ss - perm_within
                if df1 > 0 and df2 > 0 and perm_within > 0:
                    perm_F.append((perm_between / df1) / (perm_within / df2))
            if np.isnan(pseudoF) or len(perm_F) == 0:
                pval = np.nan
            else:
                perm_F = np.array(perm_F)
                pval = (1 + (perm_F >= pseudoF).sum()) / (1 + len(perm_F))
            return {"method": "permanova_fallback", "group": group_col, "pvalue": pval, "pseudoF": pseudoF}

    def pairwise_permanova(dist_df: pd.DataFrame, meta: pd.DataFrame, group_col: str):
        """Run pairwise PERMANOVA for each group pair; returns list of dicts."""
        results = []
        if dist_df.empty or group_col not in meta.columns:
            return results
        groups = meta.set_index("sample")[group_col].reindex(dist_df.index)
        uniq = [g for g in groups.dropna().unique()]
        for i in range(len(uniq)):
            for j in range(i + 1, len(uniq)):
                g1, g2 = uniq[i], uniq[j]
                idx = groups[groups.isin([g1, g2])].index
                if len(idx) < 3:
                    results.append({"group": group_col, "g1": g1, "g2": g2, "pvalue": np.nan, "pseudoF": np.nan, "method": "permanova", "n": len(idx)})
                    continue
                sub_dist = dist_df.loc[idx, idx]
                sub_meta = meta[meta["sample"].isin(idx)]
                res = run_permanova(sub_dist, sub_meta, group_col=group_col)
                res.update({"g1": g1, "g2": g2, "n": len(idx)})
                results.append(res)
        return results

    # ASV-level beta diversity (Bray–Curtis and Jaccard) using full ASV table
    if asv_counts:
        asv_mat = np.array(asv_counts, dtype=float)  # shape: n_asv x n_samples
        sample_sums = asv_mat.sum(axis=0, keepdims=True)
        with np.errstate(divide="ignore", invalid="ignore"):
            asv_rel = np.divide(asv_mat, sample_sums, out=np.zeros_like(asv_mat), where=sample_sums != 0)
        asv_rel_samples = asv_rel.T  # samples x asv
        bc_asv_df = _pairwise(asv_rel_samples, samples, _bray)
        asv_presence = (asv_mat > 0).astype(float).T  # samples x asv presence
        jaccard_asv_df = _pairwise(asv_presence, samples, _jaccard)
        dist_asv_long = summarize_within_between(bc_asv_df, meta_df, group_col="site")
        dist_asv_treat = summarize_within_between(bc_asv_df, meta_df, group_col="treatment")
        stats_asv = compare_within_between(dist_asv_long, "asv_site")
        stats_asv_treat = compare_within_between(dist_asv_treat, "asv_treatment")
        permanova_asv = run_permanova(bc_asv_df, meta_df, group_col="site")
        permanova_asv_treat = run_permanova(bc_asv_df, meta_df, group_col="treatment")
        pairwise_treat = pairwise_permanova(bc_asv_df, meta_df, group_col="treatment")
    else:
        bc_asv_df = pd.DataFrame()
        jaccard_asv_df = pd.DataFrame()
        dist_asv_long = pd.DataFrame()
        dist_asv_treat = pd.DataFrame()
        stats_asv = {"level": "asv_site", "within_median": np.nan, "between_median": np.nan, "pvalue": np.nan}
        stats_asv_treat = {"level": "asv_treatment", "within_median": np.nan, "between_median": np.nan, "pvalue": np.nan}
        permanova_asv = {"method": "permanova", "group": "site", "pvalue": np.nan, "pseudoF": np.nan}
        permanova_asv_treat = {"method": "permanova", "group": "treatment", "pvalue": np.nan, "pseudoF": np.nan}
        pairwise_treat = []
    stats_genus = compare_within_between(dist_genus_long, "genus_treatment")

    # NMDS analysis at genus level (optional)
    try:
        from scipy.spatial.distance import pdist, squareform
        from sklearn.manifold import MDS
        # Only use samples with non‑zero library size
        samples_nonzero = [s for s in samples if libsize[s] > 0]
        gen_mat = genus_rel.loc[samples_nonzero].values
        if len(samples_nonzero) >= 2 and gen_mat.shape[1] > 0:
            bc = squareform(pdist(gen_mat, metric="braycurtis"))
            bc = np.nan_to_num(bc, nan=0.0, posinf=1.0, neginf=0.0)
            nmds = MDS(
                n_components=2,
                dissimilarity="precomputed",
                random_state=42,
                metric=False,
                max_iter=300,
                n_init=4,
            )
            coords = nmds.fit_transform(bc)
            stress = float(nmds.stress_)
            nmds_df = pd.DataFrame(coords, columns=["NMDS1", "NMDS2"])
            nmds_df["sample"] = samples_nonzero
            nmds_df = nmds_df.merge(meta_df, on="sample", how="left")
            nmds_df["stress"] = stress
        else:
            nmds_df = pd.DataFrame(columns=["NMDS1", "NMDS2", "sample", "site", "plot", "rep", "depth", "stress"])
    except Exception as e:
        print(f"NMDS skipped due to missing deps or error: {e}")
        nmds_df = pd.DataFrame(columns=["NMDS1", "NMDS2", "sample", "site", "plot", "rep", "depth", "stress"])

    # NMDS at ASV level (optional)
    try:
        if asv_counts and len(samples) >= 2:
            from scipy.spatial.distance import pdist, squareform
            from sklearn.manifold import MDS
            asv_rel_samples = asv_rel_samples  # already computed
            bc = squareform(pdist(asv_rel_samples, metric="braycurtis"))
            bc = np.nan_to_num(bc, nan=0.0, posinf=1.0, neginf=0.0)
            nmds_asv = MDS(
                n_components=2,
                dissimilarity="precomputed",
                random_state=42,
                metric=False,
                max_iter=300,
                n_init=4,
            )
            coords = nmds_asv.fit_transform(bc)
            stress = float(nmds_asv.stress_)
            nmds_asv_df = pd.DataFrame(coords, columns=["NMDS1", "NMDS2"])
            nmds_asv_df["sample"] = samples
            nmds_asv_df = nmds_asv_df.merge(meta_df, on="sample", how="left")
            nmds_asv_df["stress"] = stress
        else:
            nmds_asv_df = pd.DataFrame(columns=["NMDS1", "NMDS2", "sample", "site", "plot", "rep", "depth", "stress"])
    except Exception as e:
        print(f"ASV NMDS skipped due to missing deps or error: {e}")
        nmds_asv_df = pd.DataFrame(columns=["NMDS1", "NMDS2", "sample", "site", "plot", "rep", "depth", "stress"])

    # Write Excel workbook
    out_xlsx = out_path / "metabarcoding_basic_analysis_FULL.xlsx"
    try:
        writer_path = out_xlsx
        with pd.ExcelWriter(writer_path, engine="openpyxl") as writer:
            meta_df.to_excel(writer, sheet_name="Sample_Metadata", index=False)
            alpha_annot.to_excel(writer, sheet_name="Alpha_Diversity", index=False)
            overall_counts.to_excel(writer, sheet_name="Overall_Taxa_Counts", index=False)
            comp_order.to_excel(writer, sheet_name="Composition_Order", index=False)
            comp_family.to_excel(writer, sheet_name="Composition_Family", index=False)
            if not comp_order_treat.empty:
                comp_order_treat.to_excel(writer, sheet_name="Composition_Order_Treatment", index=False)
            if not comp_family_treat.empty:
                comp_family_treat.to_excel(writer, sheet_name="Composition_Family_Treatment", index=False)
            comp_genus.to_excel(writer, sheet_name="Composition_Genus", index=False)
            bc_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="BrayCurtis_Genus", index=False)
            jaccard_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="Jaccard_Genus", index=False)
            nmds_df.to_excel(writer, sheet_name="NMDS_Genus", index=False)
            if not dist_genus_long.empty:
                dist_genus_long.to_excel(writer, sheet_name="Distances_Genus", index=False)
            if not bc_asv_df.empty:
                bc_asv_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="BrayCurtis_ASV", index=False)
                jaccard_asv_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="Jaccard_ASV", index=False)
                nmds_asv_df.to_excel(writer, sheet_name="NMDS_ASV", index=False)
                if not dist_asv_long.empty:
                    dist_asv_long.to_excel(writer, sheet_name="Distances_ASV", index=False)
            stats_df = pd.DataFrame([stats_genus, stats_asv, stats_asv_treat])
            stats_df.to_excel(writer, sheet_name="Distance_Stats", index=False)
        pd.DataFrame([permanova_asv, permanova_asv_treat]).to_excel(writer, sheet_name="PERMANOVA_ASV", index=False)
        if pairwise_treat:
            pd.DataFrame(pairwise_treat).to_excel(writer, sheet_name="PERMANOVA_Treatment_Pairs", index=False)
    except PermissionError:
        # If the workbook is open/locked, fall back to an alternate name
        alt_path = out_path / "metabarcoding_basic_analysis_FULL_alt.xlsx"
        writer_path = alt_path
        with pd.ExcelWriter(writer_path, engine="openpyxl") as writer:
            meta_df.to_excel(writer, sheet_name="Sample_Metadata", index=False)
            alpha_annot.to_excel(writer, sheet_name="Alpha_Diversity", index=False)
            overall_counts.to_excel(writer, sheet_name="Overall_Taxa_Counts", index=False)
            comp_order.to_excel(writer, sheet_name="Composition_Order", index=False)
            comp_family.to_excel(writer, sheet_name="Composition_Family", index=False)
            if not comp_order_treat.empty:
                comp_order_treat.to_excel(writer, sheet_name="Composition_Order_Treatment", index=False)
            if not comp_family_treat.empty:
                comp_family_treat.to_excel(writer, sheet_name="Composition_Family_Treatment", index=False)
            comp_genus.to_excel(writer, sheet_name="Composition_Genus", index=False)
            bc_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="BrayCurtis_Genus", index=False)
            jaccard_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="Jaccard_Genus", index=False)
            nmds_df.to_excel(writer, sheet_name="NMDS_Genus", index=False)
            if not dist_genus_long.empty:
                dist_genus_long.to_excel(writer, sheet_name="Distances_Genus", index=False)
            if not bc_asv_df.empty:
                bc_asv_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="BrayCurtis_ASV", index=False)
                jaccard_asv_df.reset_index().rename(columns={"index": "sample"}).to_excel(writer, sheet_name="Jaccard_ASV", index=False)
                nmds_asv_df.to_excel(writer, sheet_name="NMDS_ASV", index=False)
                if not dist_asv_long.empty:
                    dist_asv_long.to_excel(writer, sheet_name="Distances_ASV", index=False)
            stats_df = pd.DataFrame([stats_genus, stats_asv, stats_asv_treat])
            stats_df.to_excel(writer, sheet_name="Distance_Stats", index=False)
            pd.DataFrame([permanova_asv, permanova_asv_treat]).to_excel(writer, sheet_name="PERMANOVA_ASV", index=False)
        print(f"⚠️ Original Excel file locked; wrote outputs to {writer_path}")

    # Plot library sizes
    plt.figure(figsize=(max(8, len(alpha_df) * 0.2), 4))
    plt.bar(alpha_df["sample"], alpha_df["library_size"])
    plt.xticks(rotation=90)
    plt.ylabel("Reads per sample")
    plt.title("Library sizes (FULL)")
    plt.tight_layout()
    plt.savefig(out_path / "plot_library_sizes_FULL.png", dpi=180)
    plt.close()

    # ASV richness by site
    alpha_annot["site"] = alpha_annot["site"].fillna("NA")
    sites = sorted(alpha_annot["site"].unique())
    plt.figure(figsize=(max(8, len(sites) * 0.4), 4))
    data = [alpha_annot.loc[alpha_annot["site"] == s, "observed_asvs"].values for s in sites]
    plt.boxplot(data, labels=sites, showfliers=False)
    plt.ylabel("Observed ASVs")
    plt.title("ASV richness by site (FULL)")
    plt.tight_layout()
    plt.savefig(out_path / "plot_richness_by_site_FULL.png", dpi=180)
    plt.close()

    # ASV richness by treatment
    alpha_annot["treatment"] = alpha_annot["treatment"].fillna("NA")
    treatments = sorted(alpha_annot["treatment"].unique())
    plt.figure(figsize=(max(8, len(treatments) * 0.5), 4))
    data_t = [alpha_annot.loc[alpha_annot["treatment"] == t, "observed_asvs"].values for t in treatments]
    plt.boxplot(data_t, labels=treatments, showfliers=False)
    plt.ylabel("Observed ASVs")
    plt.title("ASV richness by treatment (FULL)")
    plt.tight_layout()
    plt.savefig(out_path / "plot_richness_by_treatment_FULL.png", dpi=180)
    plt.close()

    # Shannon index by site
    plt.figure(figsize=(max(8, len(sites) * 0.4), 4))
    data = [alpha_annot.loc[alpha_annot["site"] == s, "shannon"].values for s in sites]
    plt.boxplot(data, labels=sites, showfliers=False)
    plt.ylabel("Shannon (ln)")
    plt.title("Shannon index by site (FULL)")
    plt.tight_layout()
    plt.savefig(out_path / "plot_shannon_by_site_FULL.png", dpi=180)
    plt.close()

    # Shannon index by treatment
    plt.figure(figsize=(max(8, len(treatments) * 0.5), 4))
    data_t = [alpha_annot.loc[alpha_annot["treatment"] == t, "shannon"].values for t in treatments]
    plt.boxplot(data_t, labels=treatments, showfliers=False)
    plt.ylabel("Shannon (ln)")
    plt.title("Shannon index by treatment (FULL)")
    plt.tight_layout()
    plt.savefig(out_path / "plot_shannon_by_treatment_FULL.png", dpi=180)
    plt.close()

    # Composition plots by treatment (order/family, top 10)
    def _stacked_bar(df_mean: pd.DataFrame, value_cols: list, title: str, outfile: Path):
        if df_mean.empty or not value_cols:
            return
        df_plot = df_mean[["treatment"] + value_cols].copy()
        # long form
        long_df = df_plot.melt(id_vars=["treatment"], value_vars=value_cols, var_name="taxon", value_name="rel_abund")
        if long_df.empty:
            return
        taxon_means = long_df.groupby("taxon")["rel_abund"].mean().sort_values(ascending=False)
        top = taxon_means.head(10).index.tolist()
        long_df["taxon_group"] = long_df["taxon"].where(long_df["taxon"].isin(top), "Other")
        plot_df = long_df.groupby(["treatment", "taxon_group"], as_index=False)["rel_abund"].sum()
        pivot = plot_df.pivot(index="treatment", columns="taxon_group", values="rel_abund").fillna(0.0)
        plt.figure(figsize=(max(8, len(pivot) * 0.5), 5))
        bottom = np.zeros(len(pivot))
        x = np.arange(len(pivot.index))
        for col in pivot.columns:
            plt.bar(x, pivot[col].values, bottom=bottom, label=col)
            bottom += pivot[col].values
        plt.xticks(x, pivot.index, rotation=45, ha="right")
        plt.ylabel("Mean relative abundance")
        plt.title(title)
        plt.legend(bbox_to_anchor=(1.04, 1), loc="upper left")
        plt.tight_layout()
        plt.savefig(outfile, dpi=180)
        plt.close()

    # order-level by treatment
    order_cols = [c for c in comp_order_treat.columns if c not in {"treatment"}]
    _stacked_bar(comp_order_treat, order_cols, "Order composition by treatment (Top 10)", out_path / "plot_composition_order_treatment.png")
    # family-level by treatment
    fam_cols = [c for c in comp_family_treat.columns if c not in {"treatment"}]
    _stacked_bar(comp_family_treat, fam_cols, "Family composition by treatment (Top 10)", out_path / "plot_composition_family_treatment.png")

    def _plot_nmds(df_coords: pd.DataFrame, title: str, outfile: Path, group_col: str = "site"):
        plt.figure(figsize=(7, 6))
        if group_col not in df_coords.columns:
            return
        groups = sorted(df_coords[group_col].dropna().unique())
        markers = ["o", "s", "D", "^", "v", "<", ">", "P", "X", "*"]
        colors = cm.get_cmap("tab10", len(groups)).colors
        for i, g in enumerate(groups):
            sub = df_coords[df_coords[group_col] == g]
            plt.scatter(sub["NMDS1"], sub["NMDS2"], label=g, alpha=0.9, marker=markers[i % len(markers)], edgecolor="black", color=colors[i % len(colors)])
            if len(sub) >= 2:
                pts = sub[["NMDS1", "NMDS2"]].values
                cov = np.cov(pts, rowvar=False)
                vals, vecs = np.linalg.eigh(cov)
                order = vals.argsort()[::-1]
                vals, vecs = vals[order], vecs[:, order]
                theta = np.degrees(np.arctan2(*vecs[:, 0][::-1]))
                width, height = 2 * np.sqrt(vals) * 2
                ell = Ellipse(xy=pts.mean(axis=0), width=width, height=height, angle=theta, alpha=0.12, facecolor=colors[i % len(colors)], edgecolor="black", linestyle="dashed")
                plt.gca().add_patch(ell)
        plt.axhline(0, linestyle="--", linewidth=0.5)
        plt.axvline(0, linestyle="--", linewidth=0.5)
        stress_val = float(df_coords["stress"].iloc[0]) if "stress" in df_coords.columns else float("nan")
        plt.xlabel("NMDS1")
        plt.ylabel("NMDS2")
        plt.title(f"{title}, stress={stress_val:.2f}")
        plt.legend(bbox_to_anchor=(1.04, 1), loc="upper left")
        plt.tight_layout()
        plt.savefig(outfile, dpi=180)
        plt.close()

    # NMDS ASV plot, if available
    if 'nmds_asv_df' in locals() and not nmds_asv_df.empty:
        _plot_nmds(nmds_asv_df, "NMDS (ASV Bray–Curtis, FULL)", out_path / "plot_nmds_asv_FULL.png", group_col="treatment")

    # Save a short summary JSON
    # Simple correlations (Spearman) with dna_conc where available
    corrs = []
    if "dna_conc" in meta_df.columns:
        meta_w = meta_df.set_index("sample")
        for metric, series in [("library_size", library_sizes), ("observed_asvs", observed_asvs), ("shannon", shannon)]:
            try:
                from scipy.stats import spearmanr
                aligned = pd.concat([series, meta_w["dna_conc"]], axis=1).dropna()
                if len(aligned) > 2:
                    rho, p = spearmanr(aligned.iloc[:, 0], aligned.iloc[:, 1])
                    corrs.append({"metric": metric, "rho": float(rho), "pvalue": float(p), "n": len(aligned)})
            except Exception:
                continue
    # Kruskal-Wallis across treatments for richness/Shannon
    kw_results = []
    try:
        from scipy.stats import kruskal
        for metric in ["observed_asvs", "shannon"]:
            groups = [alpha_annot.loc[alpha_annot["treatment"] == t, metric].dropna().values for t in treatments if len(alpha_annot.loc[alpha_annot["treatment"] == t, metric].dropna()) > 0]
            if len(groups) >= 2:
                stat, p = kruskal(*groups)
                kw_results.append({"metric": metric, "test": "kruskal_treatment", "stat": float(stat), "pvalue": float(p), "k": len(groups)})
    except Exception:
        pass

    summary = {
        "n_samples_total": len(sample_cols),
        "n_samples_nonzero": int(sum(1 for s in sample_cols if libsize[s] > 0)),
        "n_asv_rows": total_rows - 1,
        "n_orders": int(len(set(k for s in agg_by_order.values() for k in s))),
        "n_families": int(len(set(k for s in agg_by_family.values() for k in s))),
        "n_genera": int(len(set(k for s in agg_by_genus.values() for k in s))),
        "distance_stats": {"genus": stats_genus, "asv_site": stats_asv, "asv_treatment": stats_asv_treat, "permanova_site": permanova_asv, "permanova_treatment": permanova_asv_treat, "permanova_treatment_pairs": pairwise_treat},
        "correlations": corrs,
        "kw_tests": kw_results,
    }
    (out_path / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print("Done. Outputs written to:", out_path.resolve())


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Perform metabarcoding analysis on an Excel ASV table.")
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT, help=f"Path to filtered_ASV_table.xlsx (default: {DEFAULT_INPUT})")
    parser.add_argument("--outdir", "-o", default=DEFAULT_OUTDIR, help=f"Directory to write outputs (default: {DEFAULT_OUTDIR})")
    args = parser.parse_args()
    main(args.input, args.outdir)
